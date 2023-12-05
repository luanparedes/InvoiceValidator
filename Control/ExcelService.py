import xlsxwriter
import openpyxl

from Control.SpreadsheetTypeEnum import SpreadsheetTypeEnum


class ExcelService:

    # region Properties
    generated_list = []
    access_keys = {'base': [], 'smart': [], 'receive': []}
    # endregion

    # region Static Methods
    @staticmethod
    def write_excel(self):
        # Create a new Excel file and add a worksheet.
        # workbook = xlsxwriter.Workbook(f'{self.folderPath}\\Validate{date.today()}.xlsx')
        workbook = xlsxwriter.Workbook(f'C:\\Users\\luans\\Downloads\\teste.xlsx')
        worksheet = workbook.add_worksheet()
        worksheet.name = "Planilha Base"
        # Column width
        worksheet.set_column('A:A', 100)

        # Add a bold format to use to highlight cells.
        bold = workbook.add_format({'bold': True})

        # Write some simple text.
        worksheet.write('A1', 'Hello')

        # Text with formatting.
        worksheet.write('A2', 'World', bold)

        # Write some numbers, with row/column notation.
        worksheet.write(2, 0, 123)
        worksheet.write(3, 0, 123.456)

        # Insert an image.
        worksheet.insert_image('B5', 'logo.png')

        workbook.close()

    @staticmethod
    def read_excel(self, spreadsheet, worksheet_type, spread_index):
        workbook = openpyxl.load_workbook(spreadsheet[spread_index])
        worksheet = workbook.active

        self.access_keys[worksheet_type] = [worksheet[i] for i in range(1, worksheet.max_row + 1)]
    # endregion
